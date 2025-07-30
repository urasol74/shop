#!/usr/bin/env python3
import sqlite3
import openpyxl
import os

def update_base_xlsx():
    """
    Обновляет base.xlsx, добавляя недостающие детские артикулы
    """
    # Читаем существующие артикулы из base.xlsx
    base_xlsx_path = os.path.join('data', 'base.xlsx')
    wb = openpyxl.load_workbook(base_xlsx_path)
    ws = wb.active
    
    existing_arts = set()
    for row in ws.iter_rows(min_row=2):
        if row[0].value:
            existing_arts.add(str(row[0].value).strip().upper())
    
    print(f"Существующих артикулов в base.xlsx: {len(existing_arts)}")
    
    # Читаем детские артикулы из base.db
    bot_db_path = "/home/ubuntu/bot_art/instance/base.db"
    conn = sqlite3.connect(bot_db_path)
    cursor = conn.cursor()
    
    cursor.execute("SELECT DISTINCT art FROM sklad_items WHERE art LIKE '%.K' ORDER BY art")
    kids_arts = [row[0].replace('.K', '') for row in cursor.fetchall()]
    conn.close()
    
    print(f"Детских артикулов в base.db: {len(kids_arts)}")
    
    # Находим недостающие артикулы
    missing_arts = [art for art in kids_arts if art not in existing_arts]
    print(f"Недостающих артикулов: {len(missing_arts)}")
    
    if not missing_arts:
        print("Все артикулы уже есть в base.xlsx!")
        return
    
    # Добавляем недостающие артикулы
    next_row = ws.max_row + 1
    added_count = 0
    
    for art in missing_arts:
        # Добавляем строку с артикулом и базовыми данными
        ws.cell(row=next_row, column=1, value=art)  # Артикул
        ws.cell(row=next_row, column=7, value='BEN.012')  # Бренд (детский)
        ws.cell(row=next_row, column=9, value='2025 весна-літо')  # Сезон
        ws.cell(row=next_row, column=10, value='дiвч')  # Пол (по умолчанию дiвч для детских товаров)
        next_row += 1
        added_count += 1
    
    # Сохраняем файл
    wb.save(base_xlsx_path)
    print(f"Добавлено {added_count} артикулов в base.xlsx")
    print("Файл обновлен!")

if __name__ == "__main__":
    update_base_xlsx() 