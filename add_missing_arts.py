#!/usr/bin/env python3
import sqlite3
import openpyxl
import os

def add_missing_arts():
    """
    Добавляет все недостающие артикулы из base.db в base.xlsx
    """
    # Читаем существующие артикулы из base.xlsx
    base_xlsx_path = os.path.join('data', 'base.xlsx')
    wb = openpyxl.load_workbook(base_xlsx_path)
    ws = wb.active
    
    existing_arts = set()
    for row in ws.iter_rows(min_row=2):
        if row[0].value:
            existing_arts.add(str(row[0].value).strip().upper())
    
    print(f"Существующих артикулов в base.xlsx: {len(existing_arts)}")
    
    # Читаем все артикулы из base.db
    bot_db_path = "/home/ubuntu/bot_art/instance/base.db"
    conn = sqlite3.connect(bot_db_path)
    cursor = conn.cursor()
    
    cursor.execute("SELECT DISTINCT art FROM sklad_items ORDER BY art")
    all_arts = [row[0] for row in cursor.fetchall()]
    conn.close()
    
    print(f"Всего артикулов в base.db: {len(all_arts)}")
    
    # Находим недостающие артикулы
    missing_arts = [art for art in all_arts if art not in existing_arts]
    print(f"Недостающих артикулов: {len(missing_arts)}")
    
    if not missing_arts:
        print("Все артикулы уже есть в base.xlsx!")
        return
    
    # Добавляем недостающие артикулы
    next_row = ws.max_row + 1
    added_count = 0
    
    for art in missing_arts:
        # Определяем тип товара по артикулу
        if art.endswith('.K'):
            # Детский товар
            gender = 'дит'  # По умолчанию для детских товаров
            brand = 'BEN.012'
        else:
            # Взрослый товар
            gender = 'унісекс'  # По умолчанию для взрослых товаров
            brand = 'BENETTON'
        
        # Добавляем строку с артикулом и базовыми данными
        ws.cell(row=next_row, column=1, value=art)  # Артикул
        ws.cell(row=next_row, column=7, value=brand)  # Бренд
        ws.cell(row=next_row, column=9, value='2025 весна-літо')  # Сезон
        ws.cell(row=next_row, column=10, value=gender)  # Пол
        next_row += 1
        added_count += 1
    
    # Сохраняем файл
    wb.save(base_xlsx_path)
    print(f"Добавлено {added_count} артикулов в base.xlsx")
    print("Файл обновлен!")

if __name__ == "__main__":
    add_missing_arts() 