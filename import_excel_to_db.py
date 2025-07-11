import openpyxl
from models import db, Product, Category
from app import app
import os
import re
import json

# --- Новый импорт товаров с доп. полями ---
def import_products():
    der_path = os.path.join('data', 'der.xlsx')
    base_path = os.path.join('data', 'base.xlsx')
    wb_der = openpyxl.load_workbook(der_path)
    ws_der = wb_der.active
    wb_base = openpyxl.load_workbook(base_path)
    ws_base = wb_base.active

    # Составляем map артикул -> (gender, season, brand)
    base_map = {}
    for row in ws_base.iter_rows(min_row=2):
        art = str(row[0].value).strip().upper() if row[0].value else None
        brand = row[6].value if len(row) > 6 else None   # G
        season = row[8].value if len(row) > 8 else None  # I
        gender = row[9].value if len(row) > 9 else None  # J
        if art:
            base_map[art] = {
                'brand': brand,
                'season': season,
                'gender': gender
            }
    print("base_map keys:", list(base_map.keys())[:10])  # диагностика

    count_new = 0
    count_upd = 0
    for row in ws_der.iter_rows(min_row=20):
        cell = row[0].value
        if not cell:
            continue
        cell = str(cell)
        if 'Размер:' not in cell or 'Цвет:' not in cell:
            continue

        # Парсим артикул и категорию
        parts = cell.split()
        if len(parts) < 2:
            continue
        art_raw = parts[0]
        art = art_raw.replace('.K', '').strip().upper()
        print("der art:", repr(art))  # диагностика
        category_name_ua = parts[1].replace(',', '')  # до запятой
        # Исключения: если категория 'НАБІР' или 'комплект', то считаем это 'БЮСТГАЛЬТЕР'
        if category_name_ua.strip().upper() in ["НАБІР", "КОМПЛЕКТ"]:
            category_name_ua = "БЮСТГАЛЬТЕР"

        # Далее парсим размер и цвет как раньше
        m = re.search(r'Размер:\s*([^,]+),\s*Цвет:\s*([A-Z0-9]+)', cell)
        if not m:
            continue
        size = m.group(1)
        color = m.group(2)
        # Количество, цены, скидка
        qty = row[3].value if len(row) > 3 else None
        old_price = row[5].value if len(row) > 5 else None
        new_price = row[7].value if len(row) > 7 else None
        sale = row[9].value if len(row) > 9 else None
        image = f"/static/pic/{art}.jpg"
        # Категория
        category = Category.query.filter_by(name_ua=category_name_ua).first()
        if not category:
            category = Category(name_ua=category_name_ua, name_ru=category_name_ua)
            db.session.add(category)
            db.session.commit()
        category_id = category.id
        # --- Логика поиска артикула в base_map ---
        # Сначала пробуем как есть
        base_info = base_map.get(art, {})
        gender = base_info.get('gender')
        brand = base_info.get('brand')
        season = base_info.get('season')
        # Если не найдено и артикул оканчивается на .K, пробуем без .K (для всех товаров)
        if not gender and art.endswith('.K'):
            art_nok = art.replace('.K', '')
            base_info_nok = base_map.get(art_nok, {})
            if base_info_nok:
                base_info = base_info_nok
                gender = base_info.get('gender')
                brand = base_info.get('brand')
                season = base_info.get('season')
        # Диагностика
        if not base_info:
            print(f"НЕ НАЙДЕН В base_map: '{art}' и '{art.replace('.K','')}'")
        else:
            print(f"НАЙДЕН В base_map: '{art if base_info == base_map.get(art, {}) else art.replace('.K','')}' -> {base_info}")
        # Данные из base.xlsx
        # Проверяем, есть ли уже такой товар (по art, color, size)
        product = Product.query.filter_by(art=art, color=color, size=size).first()
        if product:
            product.name = category_name_ua
            product.category_id = category_id
            product.qty = qty
            product.old_price = old_price
            product.new_price = new_price
            product.sale = sale
            product.image = image
            product.gender = gender
            product.season = season
            product.brand = brand
            count_upd += 1
        else:
            product = Product(
                art=art,  # только art без .K
                name=category_name_ua,
                category_id=category_id,
                color=color,
                size=size,
                qty=qty,
                old_price=old_price,
                new_price=new_price,
                sale=sale,
                image=image,
                gender=gender,
                season=season,
                brand=brand
            )
            db.session.add(product)
            count_new += 1
    db.session.commit()
    print(f"Добавлено новых: {count_new}, обновлено: {count_upd}")

# Загружаем словарь переводов категорий
SECTIONS_BOOK_PATH = os.path.join('static', 'sections-book.json')
with open(SECTIONS_BOOK_PATH, encoding='utf-8') as f:
    SECTIONS_BOOK = json.load(f)

def get_category_ru(name_ua):
    if not name_ua:
        return ''
    return SECTIONS_BOOK.get(name_ua.upper().strip(), name_ua)

def import_categories():
    der_path = os.path.join('data', 'der.xlsx')
    wb_der = openpyxl.load_workbook(der_path)
    ws_der = wb_der.active
    categories_ua = set()
    for row in ws_der.iter_rows(min_row=20):
        cell = row[0].value
        if not cell:
            continue
        cell = str(cell)
        # Парсим только строки с нужным форматом
        m = re.match(r"([A-Z0-9]+)\.?[A-Z]*\s+([^,]+),", cell)
        if not m:
            continue
        category_name_ua = m.group(2).strip()
        if category_name_ua:
            categories_ua.add(category_name_ua)
    # Добавляем в базу только новые категории
    count_new = 0
    for name_ua in categories_ua:
        exists = Category.query.filter_by(name_ua=name_ua).first()
        if not exists:
            name_ru = get_category_ru(name_ua)
            cat = Category(name_ua=name_ua, name_ru=name_ru)
            db.session.add(cat)
            count_new += 1
    db.session.commit()
    print(f"Добавлено новых категорий: {count_new}")

def import_additional_categories_from_base():
    base_path = os.path.join('data', 'base.xlsx')
    wb_base = openpyxl.load_workbook(base_path)
    ws_base = wb_base.active
    genders = set()
    brands = set()
    seasons = set()
    for row in ws_base.iter_rows(min_row=2):
        # gender (J)
        if len(row) > 9 and row[9].value:
            gender = str(row[9].value).strip()
            if gender:
                genders.add(gender)
        # brand (G)
        if len(row) > 6 and row[6].value:
            brand = str(row[6].value).strip()
            if brand:
                brands.add(brand)
        # season (H)
        if len(row) > 7 and row[7].value:
            season = str(row[7].value).strip()
            if season:
                seasons.add(season)
    # Добавляем в базу
    count_new = 0
    for val in genders | brands | seasons:
        if not val:
            continue
        exists = Category.query.filter_by(name_ua=val).first()
        if not exists:
            name_ru = get_category_ru(val)
            cat = Category(name_ua=val, name_ru=name_ru)
            db.session.add(cat)
            count_new += 1
    db.session.commit()
    print(f"Добавлено новых категорий из base.xlsx: {count_new}")

if __name__ == "__main__":
    with app.app_context():
        db.create_all()  # Создаёт все таблицы, если их нет
        import_categories()
        import_additional_categories_from_base()
        import_products()
        # Тестовый запрос к базе
        print("\nПервые 5 товаров из базы:")
        products = Product.query.limit(5).all()
        for p in products:
            cat = Category.query.get(p.category_id)
            print(f"art={p.art}, category={cat.name_ua if cat else None}, size={p.size}, color={p.color}, old_price={p.old_price}, new_price={p.new_price}, gender={p.gender}, brand={p.brand}, season={p.season}")