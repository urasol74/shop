import openpyxl
import sqlite3
import os

def create_base_xlsx():
    """
    Создает базовый справочный файл base.xlsx с данными о поле, сезоне и бренде
    """
    # Подключаемся к базе данных
    db_path = "instance/shop.db"
    if not os.path.exists(db_path):
        print(f"Ошибка: База данных {db_path} не найдена")
        return
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Получаем все уникальные артикулы
    cursor.execute("SELECT DISTINCT art FROM product ORDER BY art")
    articles = cursor.fetchall()
    
    print(f"Найдено {len(articles)} уникальных артикулов")
    
    # Создаем новый Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Справочник"
    
    # Заголовки
    headers = [
        'Артикул', 'Название', 'Цвет', 'Размер', 'Количество', 
        'Старая цена', 'Новая цена', 'Бренд', 'Сезон', 'Пол'
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Заполняем данными
    row = 2
    for article in articles:
        art = article[0]
        
        # Получаем данные о товаре
        cursor.execute("""
            SELECT name, color, size, qty, old_price, new_price 
            FROM product 
            WHERE art = ? 
            LIMIT 1
        """, (art,))
        
        product_data = cursor.fetchone()
        if product_data:
            name, color, size, qty, old_price, new_price = product_data
            
            # Определяем пол по артикулу (простая логика)
            gender = determine_gender(art, name)
            
            # Определяем бренд
            brand = determine_brand(art, name)
            
            # Определяем сезон
            season = determine_season(art, name)
            
            # Записываем в Excel
            ws.cell(row=row, column=1, value=art)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=color)
            ws.cell(row=row, column=4, value=size)
            ws.cell(row=row, column=5, value=qty)
            ws.cell(row=row, column=6, value=old_price)
            ws.cell(row=row, column=7, value=new_price)
            ws.cell(row=row, column=8, value=brand)
            ws.cell(row=row, column=9, value=season)
            ws.cell(row=row, column=10, value=gender)
            
            row += 1
    
    conn.close()
    
    # Сохраняем файл
    output_path = "data/base.xlsx"
    wb.save(output_path)
    print(f"Справочный файл создан: {output_path}")
    print(f"Записано {row-2} записей")

def determine_gender(art, name):
    """
    Определяет пол по артикулу и названию
    """
    art_upper = art.upper()
    name_upper = name.upper() if name else ""
    
    # Логика определения пола
    if any(keyword in name_upper for keyword in ['ЖЕН', 'ЖІН', 'ДІВЧ', 'ДЕВОЧКА']):
        return 'Жін'
    elif any(keyword in name_upper for keyword in ['ЧОЛ', 'МУЖ', 'ХЛОПЧ', 'МАЛЬЧИК']):
        return 'Чол'
    elif any(keyword in name_upper for keyword in ['ДІТИ', 'ДЕТИ', 'ДИТЯЧИЙ']):
        return 'Діти'
    else:
        # Попробуем определить по артикулу
        if any(prefix in art_upper for prefix in ['W', 'F', 'Ж']):
            return 'Жін'
        elif any(prefix in art_upper for prefix in ['M', 'Ч', 'Х']):
            return 'Чол'
        else:
            return 'Унісекс'

def determine_brand(art, name):
    """
    Определяет бренд по артикулу и названию
    """
    art_upper = art.upper()
    name_upper = name.upper() if name else ""
    
    if 'BEN' in art_upper or 'BENETTON' in name_upper:
        return 'Benetton'
    elif 'UNDERCOLOR' in art_upper or 'БЕЛЬЕ' in name_upper:
        return 'Undercolor'
    else:
        return 'Benetton'  # По умолчанию

def determine_season(art, name):
    """
    Определяет сезон по артикулу и названию
    """
    art_upper = art.upper()
    name_upper = name.upper() if name else ""
    
    if any(keyword in name_upper for keyword in ['ЛЕТО', 'ЛІТО', 'ВЕСНА']):
        return 'Весна-Літо'
    elif any(keyword in name_upper for keyword in ['ЗИМА', 'ОСЕНЬ', 'ОСІНЬ']):
        return 'Осінь-Зима'
    else:
        return 'Всесезонний'

if __name__ == "__main__":
    create_base_xlsx() 