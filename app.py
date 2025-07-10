from flask import Flask, render_template, send_from_directory, abort, request, url_for
import openpyxl
import os
import json
import shutil
import re
from flask_sqlalchemy import SQLAlchemy

# Загружаем словарь красивых названий категорий
with open(os.path.join('static', 'sections-book.json'), encoding='utf-8') as f:
    SECTIONS_BOOK = json.load(f)

def get_pretty_category(category):
    if not category:
        return ''
    return SECTIONS_BOOK.get(category.upper().strip(), category.capitalize())

app = Flask(__name__)

# Глобальная переменная для хранения категорий
kids_categories = []

# Конфигурация SQLite
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///shop.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name_ua = db.Column(db.String, unique=True, nullable=False)
    name_ru = db.Column(db.String, nullable=False)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    art = db.Column(db.String, nullable=False)
    name = db.Column(db.String, nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'))
    color = db.Column(db.String)
    size = db.Column(db.String)
    qty = db.Column(db.Float)
    old_price = db.Column(db.Float)
    new_price = db.Column(db.Float)
    sale = db.Column(db.String)
    image = db.Column(db.String)
    gender = db.Column(db.String)
    section = db.Column(db.String)
    category = db.relationship('Category', backref='products')

def load_kids_categories():
    xlsx_path = os.path.join('data', 'der.xlsx')
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    categories = set()
    for row in ws.iter_rows(min_row=1):
        values = [str(cell.value) if cell.value is not None else "" for cell in row[:3]]
        combined = " ".join(values).strip()
        if not combined:
            continue
        parts = combined.split()
        if len(parts) < 2:
            continue
        article = parts[0]
        if ".K" in article:
            category = parts[1]
            categories.add(category)
    categories = sorted(categories)
    # Сохраняем в kids_sections.json
    with open('kids_sections.json', 'w', encoding='utf-8') as f:
        json.dump(categories, f, ensure_ascii=False, indent=2)
    return categories

# Загружаем категории один раз при запуске
kids_categories = load_kids_categories()

# Копируем файл Новый.xlsx в /home/ubuntu/shop/data/der.xlsx при запуске
src = "/home/ubuntu/bot_art/data/Новый.xlsx"
dst = "/home/ubuntu/shop/data/der.xlsx"
os.makedirs(os.path.dirname(dst), exist_ok=True)
shutil.copyfile(src, dst)

# --- Автоматическое обновление базы данных из der.xlsx ---
# import subprocess
# subprocess.run(['python3', 'import_excel_to_db.py'])

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/kids')
def kids():
    categories = [{"raw": cat, "pretty": get_pretty_category(cat)} for cat in get_kids_categories()]
    return render_template(
        'sections.html',
        section_verbose='Дети',
        categories=categories,
        section_url='kids_section',
        back_url='/'
    )

def get_gender_map():
    base_path = os.path.join('data', 'base.xlsx')
    wb = openpyxl.load_workbook(base_path)
    ws = wb.active
    gender_map = {}
    for row in ws.iter_rows(min_row=2):
        art = str(row[0].value).replace('.0', '').replace('.K', '').strip().upper() if row[0].value else None
        gender_cell = row[9].value if len(row) > 9 else None
        if art and gender_cell:
            if 'дiвч' in str(gender_cell).lower():
                gender = 'Девочка'
            elif 'дит' in str(gender_cell).lower():
                gender = 'Детское'
            elif 'хлопч' in str(gender_cell).lower():
                gender = 'Мальчик'
            else:
                gender = None
            gender_map[art] = gender
    return gender_map

gender_map = get_gender_map()

@app.route('/kids_category/<category>')
def kids_category(category):
    # Получаем товары для категории (пример: как в kids_section)
    section_key = category.upper().strip()
    products = kids_products_by_section.get(section_key, [])
    # Добавляем поле gender для каждого товара
    for p in products:
        art_key = p['art'].replace('.K', '').strip().upper()
        p['gender'] = gender_map.get(art_key)
    return render_template('kids_category.html', category=category, products=products)

def get_kids_categories():
    wb = openpyxl.load_workbook(dst)
    ws = wb.active
    categories = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell = row[0]
        if not cell:
            continue
        cell = str(cell)
        # Проверяем наличие .K (kids)
        if ".K" in cell:
            after_k = cell.split(".K", 1)[1].strip()
            if "," in after_k:
                section = after_k.split(",", 1)[0].strip()
            else:
                section = after_k.strip()
            if section:
                categories.add(section)
    return sorted(categories)

def build_kids_products():
    wb = openpyxl.load_workbook(dst)
    ws = wb.active
    products_by_section = {}
    all_rows = list(ws.iter_rows(min_row=2))
    for row in all_rows:
        # --- Фильтрация по количеству ---
        qty = row[3].value if len(row) > 3 else None
        if not qty or str(qty).strip() == "0" or "-" in str(qty):
            continue
        cell = row[0].value
        if not cell:
            continue
        cell = str(cell)
        if ".K" in cell:
            after_k = cell.split(".K", 1)[1].strip()
            if "," in after_k:
                section = after_k.split(",", 1)[0].strip()
            else:
                section = after_k.strip()
            section_key = section.upper().strip()
            art = cell.split(".K", 1)[0].strip() + ".K"
            art_no_k = art.replace('.K', '')
            # --- Уникальность по артикулу ---
            if section_key not in products_by_section:
                products_by_section[section_key] = []
                seen_arts = set()
            else:
                seen_arts = {p['art'] for p in products_by_section[section_key]}
            if art in seen_arts:
                continue
            name = section
            old_price = row[5].value if len(row) > 5 else None
            new_price = row[7].value if len(row) > 7 else None
            sale = row[9].value if len(row) > 9 else None
            # --- Форматирование sale ---
            sale_str = None
            try:
                if sale is not None:
                    sale_num = float(sale)
                    sale_rounded = int(round(sale_num / 10.0) * 10)
                    sale_str = f"-{sale_rounded}%"
            except Exception:
                sale_str = None
            image = f"/static/pic/{art_no_k}.jpg"
            # --- Собираем все строки с этим артикулом для meta-block ---
            color_size_map = {}
            for r in all_rows:
                c = r[0].value
                if not c:
                    continue
                c = str(c)
                if art in c:
                    # Парсим только из первой ячейки строки
                    color = None
                    size = None
                    if 'Цвет:' in c:
                        color = c.split('Цвет:')[-1].split(',')[0].strip()
                    if 'Размер:' in c:
                        size = c.split('Размер:')[-1].split(',')[0].strip()
                    if color:
                        if color not in color_size_map:
                            color_size_map[color] = []
                        if size and size not in color_size_map[color]:
                            color_size_map[color].append(size)
            color_size_list = [{"color": k, "sizes": v} for k, v in color_size_map.items()]
            product = {
                "name": name,
                "art": art,
                "old_price": old_price,
                "new_price": new_price,
                "sale": sale_str,
                "image": image,
                "color_size_map": color_size_list
            }
            products_by_section[section_key].append(product)
    with open("kids_products.json", "w", encoding="utf-8") as f:
        json.dump(products_by_section, f, ensure_ascii=False, indent=2)
    return products_by_section

# При запуске сервера
kids_products_by_section = build_kids_products()

@app.route('/kids/<section>')
def kids_section(section):
    section_key = section.upper().strip()
    products = kids_products_by_section.get(section_key, [])
    # Добавляем поле gender для каждого товара
    for p in products:
        art_key = p['art'].replace('.K', '').strip().upper()
        p['gender'] = gender_map.get(art_key)
    # Определяем, какие гендеры реально есть в этой категории
    available_genders = set()
    for p in products:
        if p.get('gender'):
            available_genders.add(p['gender'])
    gender_order = ['Мальчик', 'Девочка', 'Детское']
    genders = [g for g in gender_order if g in available_genders]
    return render_template('products.html', section='kids', section_verbose='Дети', category=section, pretty_category=get_pretty_category(section), products=products, genders=genders, back_url='/kids')

@app.route('/kids/<section>/<art>')
def kids_product(section, art):
    section_key = section.upper().strip()
    products = kids_products_by_section.get(section_key, [])
    product = next((p for p in products if p['art'] == art), None)
    return render_template('product.html', product=product, section_verbose='Дети', back_url=f'/kids/{section}')

# Заглушки для men и woman
# def get_men_categories():http://178.212.198.23/
#     return []
# def get_woman_categories():
#     return []

@app.route('/.well-known/acme-challenge/<path:filename>')
def letsencrypt_challenge(filename):
    acme_dir = os.path.join(app.root_path, '.well-known', 'acme-challenge')
    return send_from_directory(acme_dir, filename)

def load_woman_categories():
    der_path = os.path.join('data', 'der.xlsx')
    base_path = os.path.join('data', 'base.xlsx')
    wb_der = openpyxl.load_workbook(der_path)
    ws_der = wb_der.active
    wb_base = openpyxl.load_workbook(base_path)
    ws_base = wb_base.active
    # Собираем все артикула с жiн из base.xlsx, но не UNDERCOLOR (теперь G)
    woman_arts = set()
    for row in ws_base.iter_rows(min_row=2):
        art = str(row[0].value).strip().upper() if row[0].value else None
        col_g = row[6].value if len(row) > 6 else None
        gender_cell = row[9].value if len(row) > 9 else None
        if art and gender_cell and 'жiн' in str(gender_cell).lower():
            if col_g and str(col_g).strip().upper() == 'UNDERCOLOR':
                continue
            woman_arts.add(art)
    # Собираем категории из der.xlsx
    categories = set()
    for row in ws_der.iter_rows(min_row=2):
        values = [str(cell.value) if cell.value is not None else "" for cell in row[:3]]
        combined = " ".join(values).strip()
        if not combined:
            continue
        parts = combined.split()
        if len(parts) < 2:
            continue
        art = parts[0].strip().upper()
        category = parts[1].strip().rstrip(',')
        if art in woman_arts:
            categories.add(category)
    categories = sorted(categories)
    with open('woman_sections.json', 'w', encoding='utf-8') as f:
        json.dump(categories, f, ensure_ascii=False, indent=2)
    return categories

woman_categories = load_woman_categories()

@app.route('/women')
def women():
    categories = [{"raw": cat, "pretty": get_pretty_category(cat)} for cat in woman_categories]
    return render_template(
        'sections.html',
        section_verbose='Женщина',
        categories=categories,
        section_url='women_section',
        back_url='/'
    )

@app.route('/men')
def men():
    categories = [{"raw": cat, "pretty": get_pretty_category(cat)} for cat in men_categories]
    return render_template(
        'sections.html',
        section_verbose='Мужчина',
        categories=categories,
        section_url='men_section',
        back_url='/'
    )

@app.route('/underwear')
def underwear():
    return render_template('underwear.html')

def build_woman_products():
    der_path = os.path.join('data', 'der.xlsx')
    base_path = os.path.join('data', 'base.xlsx')
    wb_der = openpyxl.load_workbook(der_path)
    ws_der = wb_der.active
    wb_base = openpyxl.load_workbook(base_path)
    ws_base = wb_base.active
    # Собираем все артикула с жiн из base.xlsx, но не UNDERCOLOR (G)
    woman_arts = set()
    for row in ws_base.iter_rows(min_row=2):
        art = str(row[0].value).strip().upper() if row[0].value else None
        col_g = row[6].value if len(row) > 6 else None
        gender_cell = row[9].value if len(row) > 9 else None
        if art and gender_cell and 'жiн' in str(gender_cell).lower():
            if col_g and str(col_g).strip().upper() == 'UNDERCOLOR':
                continue
            woman_arts.add(art)
    # Собираем товары по категориям
    products_by_section = {}
    all_rows = list(ws_der.iter_rows(min_row=2))
    for row in all_rows:
        # --- Фильтрация по количеству ---
        qty = row[3].value if len(row) > 3 else None
        if not qty or str(qty).strip() == "0" or "-" in str(qty):
            continue
        values = [str(cell.value) if cell.value is not None else "" for cell in row[:3]]
        combined = " ".join(values).strip()
        if not combined:
            continue
        parts = combined.split()
        if len(parts) < 2:
            continue
        art = parts[0].strip().upper()
        category = parts[1].strip().rstrip(',')
        if art not in woman_arts:
            continue
        section_key = category.upper().strip()
        # --- Уникальность по артикулу ---
        if section_key not in products_by_section:
            products_by_section[section_key] = []
            seen_arts = set()
        else:
            seen_arts = {p['art'] for p in products_by_section[section_key]}
        if art in seen_arts:
            continue
        name = category
        old_price = row[5].value if len(row) > 5 else None
        new_price = row[7].value if len(row) > 7 else None
        sale = row[9].value if len(row) > 9 else None
        # --- Форматирование sale ---
        sale_str = None
        try:
            if sale is not None:
                sale_num = float(sale)
                sale_rounded = int(round(sale_num / 10.0) * 10)
                sale_str = f"-{sale_rounded}%"
        except Exception:
            sale_str = None
        image = f"/static/pic/{art}.jpg"
        # --- Собираем все строки с этим артикулом для meta-block ---
        color_size_map = {}
        for r in all_rows:
            c = r[0].value
            if not c:
                continue
            c = str(c)
            if art in c:
                color = None
                size = None
                if 'Цвет:' in c:
                    color = c.split('Цвет:')[-1].split(',')[0].strip()
                if 'Размер:' in c:
                    size = c.split('Размер:')[-1].split(',')[0].strip()
                if color:
                    if color not in color_size_map:
                        color_size_map[color] = []
                    if size and size not in color_size_map[color]:
                        color_size_map[color].append(size)
        color_size_list = [{"color": k, "sizes": v} for k, v in color_size_map.items()]
        product = {
            "name": name,
            "art": art,
            "old_price": old_price,
            "new_price": new_price,
            "sale": sale_str,
            "image": image,
            "color_size_map": color_size_list
        }
        products_by_section[section_key].append(product)
    with open("woman_products.json", "w", encoding="utf-8") as f:
        json.dump(products_by_section, f, ensure_ascii=False, indent=2)
    return products_by_section

woman_products_by_section = build_woman_products()

@app.route('/women/<section>')
def women_section(section):
    section_key = section.upper().strip()
    products = woman_products_by_section.get(section_key, [])
    return render_template('products.html', section='women', section_verbose='Женщина', category=section, pretty_category=get_pretty_category(section), products=products, back_url='/women')

@app.route('/women/<section>/<art>')
def women_product(section, art):
    section_key = section.upper().strip()
    products = woman_products_by_section.get(section_key, [])
    product = next((p for p in products if p['art'] == art), None)
    return render_template('product.html', product=product, section_verbose='Женщина', back_url=f'/women/{section}')

def load_men_categories():
    der_path = os.path.join('data', 'der.xlsx')
    base_path = os.path.join('data', 'base.xlsx')
    wb_der = openpyxl.load_workbook(der_path)
    ws_der = wb_der.active
    wb_base = openpyxl.load_workbook(base_path)
    ws_base = wb_base.active
    # Собираем все артикула с чол из base.xlsx, но не UNDERCOLOR (G)
    men_arts = set()
    for row in ws_base.iter_rows(min_row=2):
        art = str(row[0].value).strip().upper() if row[0].value else None
        col_g = row[6].value if len(row) > 6 else None
        gender_cell = row[9].value if len(row) > 9 else None
        if art and gender_cell and 'чол' in str(gender_cell).lower():
            if col_g and str(col_g).strip().upper() == 'UNDERCOLOR':
                continue
            men_arts.add(art)
    # Собираем категории из der.xlsx
    categories = set()
    for row in ws_der.iter_rows(min_row=2):
        values = [str(cell.value) if cell.value is not None else "" for cell in row[:3]]
        combined = " ".join(values).strip()
        if not combined:
            continue
        parts = combined.split()
        if len(parts) < 2:
            continue
        art = parts[0].strip().upper()
        category = parts[1].strip().rstrip(',')
        if art in men_arts:
            categories.add(category)
    categories = sorted(categories)
    with open('men_sections.json', 'w', encoding='utf-8') as f:
        json.dump(categories, f, ensure_ascii=False, indent=2)
    return categories

men_categories = load_men_categories()

def build_men_products():
    der_path = os.path.join('data', 'der.xlsx')
    base_path = os.path.join('data', 'base.xlsx')
    wb_der = openpyxl.load_workbook(der_path)
    ws_der = wb_der.active
    wb_base = openpyxl.load_workbook(base_path)
    ws_base = wb_base.active
    # Собираем все артикула с чол из base.xlsx, но не UNDERCOLOR (G)
    men_arts = set()
    for row in ws_base.iter_rows(min_row=2):
        art = str(row[0].value).strip().upper() if row[0].value else None
        col_g = row[6].value if len(row) > 6 else None
        gender_cell = row[9].value if len(row) > 9 else None
        if art and gender_cell and 'чол' in str(gender_cell).lower():
            if col_g and str(col_g).strip().upper() == 'UNDERCOLOR':
                continue
            men_arts.add(art)
    # Собираем товары по категориям
    products_by_section = {}
    all_rows = list(ws_der.iter_rows(min_row=2))
    for row in all_rows:
        # --- Фильтрация по количеству ---
        qty = row[3].value if len(row) > 3 else None
        if not qty or str(qty).strip() == "0" or "-" in str(qty):
            continue
        values = [str(cell.value) if cell.value is not None else "" for cell in row[:3]]
        combined = " ".join(values).strip()
        if not combined:
            continue
        parts = combined.split()
        if len(parts) < 2:
            continue
        art = parts[0].strip().upper()
        category = parts[1].strip().rstrip(',')
        if art not in men_arts:
            continue
        section_key = category.upper().strip()
        # --- Уникальность по артикулу ---
        if section_key not in products_by_section:
            products_by_section[section_key] = []
            seen_arts = set()
        else:
            seen_arts = {p['art'] for p in products_by_section[section_key]}
        if art in seen_arts:
            continue
        name = category
        old_price = row[5].value if len(row) > 5 else None
        new_price = row[7].value if len(row) > 7 else None
        sale = row[9].value if len(row) > 9 else None
        # --- Форматирование sale ---
        sale_str = None
        try:
            if sale is not None:
                sale_num = float(sale)
                sale_rounded = int(round(sale_num / 10.0) * 10)
                sale_str = f"-{sale_rounded}%"
        except Exception:
            sale_str = None
        image = f"/static/pic/{art}.jpg"
        # --- Собираем все строки с этим артикулом для meta-block ---
        color_size_map = {}
        for r in all_rows:
            c = r[0].value
            if not c:
                continue
            c = str(c)
            if art in c:
                color = None
                size = None
                if 'Цвет:' in c:
                    color = c.split('Цвет:')[-1].split(',')[0].strip()
                if 'Размер:' in c:
                    size = c.split('Размер:')[-1].split(',')[0].strip()
                if color:
                    if color not in color_size_map:
                        color_size_map[color] = []
                    if size and size not in color_size_map[color]:
                        color_size_map[color].append(size)
        color_size_list = [{"color": k, "sizes": v} for k, v in color_size_map.items()]
        product = {
            "name": name,
            "art": art,
            "old_price": old_price,
            "new_price": new_price,
            "sale": sale_str,
            "image": image,
            "color_size_map": color_size_list
        }
        products_by_section[section_key].append(product)
    with open("men_products.json", "w", encoding="utf-8") as f:
        json.dump(products_by_section, f, ensure_ascii=False, indent=2)
    return products_by_section

men_products_by_section = build_men_products()

@app.route('/men/<section>')
def men_section(section):
    section_key = section.upper().strip()
    products = men_products_by_section.get(section_key, [])
    return render_template('products.html', section='men', section_verbose='Мужчина', category=section, pretty_category=get_pretty_category(section), products=products, back_url='/men')

@app.route('/men/<section>/<art>')
def men_product(section, art):
    section_key = section.upper().strip()
    products = men_products_by_section.get(section_key, [])
    product = next((p for p in products if p['art'] == art), None)
    return render_template('product.html', product=product, section_verbose='Мужчина', back_url=f'/men/{section}')

def build_underwear_sections_and_products():
    der_path = os.path.join('data', 'der.xlsx')
    base_path = os.path.join('data', 'base.xlsx')
    wb_der = openpyxl.load_workbook(der_path)
    ws_der = wb_der.active
    wb_base = openpyxl.load_workbook(base_path)
    ws_base = wb_base.active
    # Собираем артикула UNDERCOLOR по группам gender
    gender_map = {
        'woman': {'j': 'жiн', 'sections': set(), 'products': {}},
        'men': {'j': 'чол', 'sections': set(), 'products': {}},
        'girl': {'j': 'дiвч', 'sections': set(), 'products': {}},
        'boy': {'j': 'хлопч', 'sections': set(), 'products': {}}
    }
    # Сначала строим map art -> (gender, is_undercolor)
    art_gender = {}
    for row in ws_base.iter_rows(min_row=2):
        art = str(row[0].value).strip().upper() if row[0].value else None
        col_g = row[6].value if len(row) > 6 else None
        gender_cell = row[9].value if len(row) > 9 else None
        if art and col_g and str(col_g).strip().upper() == 'UNDERCOLOR' and gender_cell:
            for key, val in gender_map.items():
                if val['j'] in str(gender_cell).lower():
                    art_gender[art] = key
    # Теперь собираем категории и товары по der.xlsx
    all_rows = list(ws_der.iter_rows(min_row=2))
    for row in all_rows:
        # --- Фильтрация по количеству ---
        qty = row[3].value if len(row) > 3 else None
        if not qty or str(qty).strip() == "0" or "-" in str(qty):
            continue
        values = [str(cell.value) if cell.value is not None else "" for cell in row[:3]]
        combined = " ".join(values).strip()
        if not combined:
            continue
        parts = combined.split()
        if len(parts) < 2:
            continue
        art = parts[0].strip().upper()
        category = parts[1].strip().rstrip(',')
        group = art_gender.get(art)
        if not group:
            continue
        section_key = category.upper().strip()
        gender_map[group]['sections'].add(category)
        # --- Уникальность по артикулу ---
        if section_key not in gender_map[group]['products']:
            gender_map[group]['products'][section_key] = []
            seen_arts = set()
        else:
            seen_arts = {p['art'] for p in gender_map[group]['products'][section_key]}
        if art in seen_arts:
            continue
        name = f"{category} {art}"
        old_price = row[5].value if len(row) > 5 else None
        new_price = row[7].value if len(row) > 7 else None
        sale = row[9].value if len(row) > 9 else None
        sale_str = None
        try:
            if sale is not None:
                sale_num = float(sale)
                sale_rounded = int(round(sale_num / 10.0) * 10)
                sale_str = f"-{sale_rounded}%"
        except Exception:
            sale_str = None
        image = f"/static/pic/{art}.jpg"
        color_size_map = {}
        for r in all_rows:
            c = r[0].value
            if not c:
                continue
            c = str(c)
            if art in c:
                color = None
                size = None
                if 'Цвет:' in c:
                    color = c.split('Цвет:')[-1].split(',')[0].strip()
                if 'Размер:' in c:
                    size = c.split('Размер:')[-1].split(',')[0].strip()
                if color:
                    if color not in color_size_map:
                        color_size_map[color] = []
                    if size and size not in color_size_map[color]:
                        color_size_map[color].append(size)
        color_size_list = [{"color": k, "sizes": v} for k, v in color_size_map.items()]
        product = {
            "name": name,
            "art": art,
            "old_price": old_price,
            "new_price": new_price,
            "sale": sale_str,
            "image": image,
            "color_size_map": color_size_list
        }
        gender_map[group]['products'][section_key].append(product)
    # Сохраняем json для каждой группы
    for key in gender_map:
        with open(f'underwear_{key}_sections.json', 'w', encoding='utf-8') as f:
            json.dump(sorted(gender_map[key]['sections']), f, ensure_ascii=False, indent=2)
        with open(f'underwear_{key}_products.json', 'w', encoding='utf-8') as f:
            json.dump(gender_map[key]['products'], f, ensure_ascii=False, indent=2)
    return gender_map

underwear_data = build_underwear_sections_and_products()

@app.route('/underwear-woman')
def underwear_woman():
    categories = [{"raw": cat, "pretty": get_pretty_category(cat)} for cat in sorted(underwear_data['woman']['sections'])]
    return render_template(
        'sections.html',
        section_verbose='Бельё — Женщина',
        categories=categories,
        section_url='underwear_woman_products',
        back_url='/underwear'
    )

@app.route('/underwear-men')
def underwear_men():
    categories = [{"raw": cat, "pretty": get_pretty_category(cat)} for cat in sorted(underwear_data['men']['sections'])]
    return render_template(
        'sections.html',
        section_verbose='Бельё — Мужчина',
        categories=categories,
        section_url='underwear_men_products',
        back_url='/underwear'
    )

@app.route('/underwear-boy')
def underwear_boy():
    categories = [{"raw": cat, "pretty": get_pretty_category(cat)} for cat in sorted(underwear_data['boy']['sections'])]
    return render_template(
        'sections.html',
        section_verbose='Бельё — Мальчик',
        categories=categories,
        section_url='underwear_boy_products',
        back_url='/underwear'
    )

@app.route('/underwear-girl')
def underwear_girl():
    categories = [{"raw": cat, "pretty": get_pretty_category(cat)} for cat in sorted(underwear_data['girl']['sections'])]
    return render_template(
        'sections.html',
        section_verbose='Бельё — Девочка',
        categories=categories,
        section_url='underwear_girl_products',
        back_url='/underwear'
    )

@app.route('/underwear-woman/<section>')
def underwear_woman_products(section):
    section_key = section.upper().strip()
    products = underwear_data['woman']['products'].get(section_key, [])
    return render_template('products.html', section='underwear-woman', section_verbose='Бельё — Женщина', category=section, pretty_category=get_pretty_category(section), products=products, back_url='/underwear-woman')

@app.route('/underwear-men/<section>')
def underwear_men_products(section):
    section_key = section.upper().strip()
    products = underwear_data['men']['products'].get(section_key, [])
    return render_template('products.html', section='underwear-men', section_verbose='Бельё — Мужчина', category=section, pretty_category=get_pretty_category(section), products=products, back_url='/underwear-men')

@app.route('/underwear-boy/<section>')
def underwear_boy_products(section):
    section_key = section.upper().strip()
    products = underwear_data['boy']['products'].get(section_key, [])
    return render_template('products.html', section='underwear-boy', section_verbose='Бельё — Мальчик', category=section, pretty_category=get_pretty_category(section), products=products, back_url='/underwear-boy')

@app.route('/underwear-girl/<section>')
def underwear_girl_products(section):
    section_key = section.upper().strip()
    products = underwear_data['girl']['products'].get(section_key, [])
    return render_template('products.html', section='underwear-girl', section_verbose='Бельё — Девочка', category=section, pretty_category=get_pretty_category(section), products=products, back_url='/underwear-girl')

@app.route('/underwear-woman/<section>/<art>')
def underwear_woman_product(section, art):
    section_key = section.upper().strip()
    products = underwear_data['woman']['products'].get(section_key, [])
    product = next((p for p in products if p['art'] == art), None)
    return render_template('product.html', product=product, section_verbose='Бельё — Женщина', back_url=f'/underwear-woman/{section}')

@app.route('/underwear-men/<section>/<art>')
def underwear_men_product(section, art):
    section_key = section.upper().strip()
    products = underwear_data['men']['products'].get(section_key, [])
    product = next((p for p in products if p['art'] == art), None)
    return render_template('product.html', product=product, section_verbose='Бельё — Мужчина', back_url=f'/underwear-men/{section}')

@app.route('/underwear-boy/<section>/<art>')
def underwear_boy_product(section, art):
    section_key = section.upper().strip()
    products = underwear_data['boy']['products'].get(section_key, [])
    product = next((p for p in products if p['art'] == art), None)
    return render_template('product.html', product=product, section_verbose='Бельё — Мальчик', back_url=f'/underwear-boy/{section}')

@app.route('/underwear-girl/<section>/<art>')
def underwear_girl_product(section, art):
    section_key = section.upper().strip()
    products = underwear_data['girl']['products'].get(section_key, [])
    product = next((p for p in products if p['art'] == art), None)
    return render_template('product.html', product=product, section_verbose='Бельё — Девочка', back_url=f'/underwear-girl/{section}')

# Универсальный роут для страницы товара
@app.route('/<section>/<category>/<art>')
def product_page(section, category, art):
    # Определяем источник данных и verbose-название раздела
    section_map = {
        'women': {
            'products': build_woman_products(),
            'verbose': 'Женщина',
            'back_url': f'/women/{category}'
        },
        'men': {
            'products': build_men_products(),
            'verbose': 'Мужчина',
            'back_url': f'/men/{category}'
        },
        'kids': {
            'products': kids_products_by_section,
            'verbose': 'Дети',
            'back_url': f'/kids/{category}'
        },
        'underwear-woman': {
            'products': underwear_woman_products,
            'verbose': 'Бельё — Женщина',
            'back_url': f'/underwear-woman/{category}'
        },
        'underwear-men': {
            'products': underwear_men_products,
            'verbose': 'Бельё — Мужчина',
            'back_url': f'/underwear-men/{category}'
        },
        'underwear-boy': {
            'products': underwear_boy_products,
            'verbose': 'Бельё — Мальчик',
            'back_url': f'/underwear-boy/{category}'
        },
        'underwear-girl': {
            'products': underwear_girl_products,
            'verbose': 'Бельё — Девочка',
            'back_url': f'/underwear-girl/{category}'
        },
    }
    if section not in section_map:
        abort(404)
    products_by_section = section_map[section]['products']
    section_verbose = section_map[section]['verbose']
    back_url = section_map[section]['back_url']
    # Ищем товар по артикулу
    products = products_by_section.get(category.upper().strip(), [])
    product = next((p for p in products if p['art'] == art), None)
    if not product:
        abort(404)
    return render_template('product.html', product=product, section_verbose=section_verbose, back_url=back_url)

@app.route('/search')
def search():
    query = request.args.get('q', '').strip().lower()
    if not query:
        return render_template('search.html', query=query, results=[], count=0)
    xlsx_path = os.path.join('data', 'der.xlsx')
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    results = []
    query_parts = query.split()
    for row in ws.iter_rows(min_row=2, values_only=True):
        cell = str(row[0]).strip() if row[0] else ""
        cell_l = cell.lower()
        # Парсим артикул, категорию, размер, цвет
        art = ''
        category = ''
        size = ''
        color = ''
        m = re.match(r"([A-Z0-9]+)\.?[A-Z]*\s+([^,]+),\s*Размер:\s*([^,]+),\s*Цвет:\s*([A-Z0-9]+)", cell)
        if m:
            art = m.group(1)
            category = m.group(2)
            size = m.group(3)
            color = m.group(4)
        # Проверяем, что все части запроса есть в нужных полях
        found = True
        for part in query_parts:
            if not (part in art.lower() or part in category.lower() or part in size.lower() or part in color.lower() or part in cell_l):
                found = False
                break
        if found:
            # --- Генерируем url на страницу товара ---
            url = "#"
            section = None
            cat_key = category.upper().strip()
            if cat_key in men_products_by_section:
                section = 'men'
            elif cat_key in woman_products_by_section:
                section = 'women'
            if section and art and category:
                url = url_for('product_page', section=section, category=cat_key, art=art)
            results.append({
                "art": art,
                "name": category,
                "color": color,
                "size": size,
                "qty": "",
                "category": category,
                "section": section or "",
                "url": url,
            })
    return render_template('search.html', query=query, results=results, count=len(results))

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000, debug=True)
