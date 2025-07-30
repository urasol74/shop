#!/usr/bin/env python3
import sqlite3
import openpyxl
import os

def fix_base_xlsx_k():
    """
    Исправляет base.xlsx - артикулы с .K должны быть записаны с .K
    """
    # Читаем все артикулы из base.db
    bot_db_path = "/home/ubuntu/bot_art/instance/base.db"
    conn = sqlite3.connect(bot_db_path)
    cursor = conn.cursor()
    
    cursor.execute("SELECT DISTINCT art FROM sklad_items ORDER BY art")
    all_arts = [row[0] for row in cursor.fetchall()]
    conn.close()
    
    print(f"Всего артикулов в base.db: {len(all_arts)}")
    
    # Читаем существующие артикулы из base.xlsx
    base_xlsx_path = os.path.join('data', 'base.xlsx')
    wb = openpyxl.load_workbook(base_xlsx_path)
    ws = wb.active
    
    existing_arts = set()
    for row in ws.iter_rows(min_row=2):
        if row[0].value:
            existing_arts.add(str(row[0].value).strip().upper())
    
    print(f"Существующих артикулов в base.xlsx: {len(existing_arts)}")
    
    # Находим артикулы с .K, которые в base.xlsx записаны без .K
    arts_to_fix = []
    for art in all_arts:
        if art.endswith('.K'):
            art_without_k = art.replace('.K', '')
            if art_without_k in existing_arts and art not in existing_arts:
                arts_to_fix.append((art_without_k, art))
    
    print(f"Артикулов для исправления: {len(arts_to_fix)}")
    
    if not arts_to_fix:
        print("Все артикулы уже правильные!")
        return
    
    # Исправляем артикулы в base.xlsx
    fixed_count = 0
    for row in ws.iter_rows(min_row=2):
        if row[0].value:
            art = str(row[0].value).strip().upper()
            for old_art, new_art in arts_to_fix:
                if art == old_art:
                    row[0].value = new_art
                    fixed_count += 1
                    break
    
    # Сохраняем файл
    wb.save(base_xlsx_path)
    print(f"Исправлено {fixed_count} артикулов в base.xlsx")
    print("Файл обновлен!")

if __name__ == "__main__":
    fix_base_xlsx_k() 